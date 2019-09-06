#Kyle Share
import openpyxl
import os

#MAC filepath
#filepath = os.path.join('/Users', 'KyleShare', 'Programming', 'caravan', 'DICKS.XLSX' )
#description_path = os.path.join('/Users', 'KyleShare', 'Programming', 'caravan', 'Item Description.xlsx')

#WINDOWS filepath
filepath = os.path.join('C:\\', 'Users', 'CaravanArms', 'Desktop', 'DICKS.XLSX' )
description_path = os.path.join('C:\\', 'Users', 'CaravanArms', 'Desktop', 'Item Description.xlsx')

#Get workbook from filepath
wb = openpyxl.load_workbook(filepath)
description_wb = openpyxl.load_workbook(description_path)

#Get all sheetnames
sheet_names = wb.sheetnames
description_sheet_names = description_wb.sheetnames

#Get first worksheet, since wb.active is set to 0 by default
first_sheet = wb.active
description_sheet = description_wb.active

#Create new workbook
new_wb = openpyxl.Workbook()
new_first_sheet = new_wb.active

#Create dictionary of (ID: item description) pairs
def create_description_dict():
    description_list = []
    #Iterate through description excel sheet
    for row_num in range(2, description_sheet.max_row + 1):
        key = str((description_sheet.cell(row=row_num, column=1).value))
        description = (description_sheet.cell(row=row_num, column=2).value)
        #Create tuples, append them to description list
        tupl = (key, description)
        description_list.append(tupl)
        #Make description dictionary global so item desctiption function can access it
        global description_dict
        description_dict = dict(description_list)

def create_carrier_dict():
    carrier_list = []
    #Iterate through prices excel sheet
    for row_num in range(2, description_sheet.max_row + 1):
        key = str((description_sheet.cell(row=row_num, column=10).value))
        #Walmart price is column 3
        carrier = (description_sheet.cell(row=row_num, column=11).value)
        #Create tuples, append them to description list
        tupl = (key, carrier)
        carrier_list.append(tupl)
        global carrier_dict
        carrier_dict = dict(carrier_list)

#Dicks needs extra dict since item number is not provided
#Function maps UPC code (provided) to corresponding item number
def create_item_num_dict():
    item_num_list = []
    #Iterate through UPC code excel sheet
    for row_num in range(2, description_sheet.max_row + 1):
        #key is upc code
        key = str((description_sheet.cell(row=row_num, column=12).value))
        item_num = (description_sheet.cell(row=row_num, column=13).value)
        #Create tuples, append them to item num list
        tupl = (key, item_num)
        item_num_list.append(tupl)
        global item_num_dict
        item_num_dict = dict(item_num_list)

  

def titles():
    titles = ["ACCOUNT(SBT CODE)", "PO#", "PO LINE", "CUSTOMER NAME", "ADDRESS 1(2ND LINE)", \
    "PHONE# (3RD LINE)", "ADDRESS 2", "CARRIER", "ITEM#", "ITEM DESCRIPTION", \
    "UNIT PRICE", "QTY", "LINE TOTAL", "TERMS", "PO TOTAL"]
    title_index = 0
    for column_num in range(1, 16):
        new_first_sheet.cell(row = 1, column = column_num).value = titles[title_index]
        title_index += 1

def account(reading, writing):
    new_first_sheet.cell(row = writing, column = 1).value = 'PA5108'

def po_num(reading, writing):
    po = first_sheet.cell(row = reading, column = 1).value
    po = str(po)
    po = po.zfill(10)
    new_first_sheet.cell(row = writing, column = 2).value = po

def po_line(reading, writing):
    line = first_sheet.cell(row = reading, column = 2).value
    new_first_sheet.cell(row = writing, column = 3).value = line

def customer_name(reading, writing):
    name = first_sheet.cell(row = reading, column = 6).value
    new_first_sheet.cell(row = writing, column = 4).value = name

#<Address, N/A> or <Store #, Store Address>
def address_1(reading, writing):
    street_address = first_sheet.cell(row = reading, column = 7).value
    address_cont = first_sheet.cell(row = reading, column = 8).value

    if address_cont != "N/A":
        street_address = "{}, {}".format(street_address, address_cont)

    new_first_sheet.cell(row = writing, column = 5).value = street_address

def phone_num(reading, writing):
    phone = first_sheet.cell(row = reading, column = 15).value
    new_first_sheet.cell(row = writing, column = 6).value = phone

#<City, State Zip>
#Initial 0's omitted on downloads, make Zip Code 5 digits
def address_2(reading, writing):
    city = first_sheet.cell(row = reading, column = 9).value
    state = first_sheet.cell(row = reading, column = 10).value

    zip_code = first_sheet.cell(row = reading, column = 11).value
    zip_code = str(zip_code)
    zip_code = zip_code.zfill(5)

    address2 = "{}, {} {}".format(city, state, zip_code)
    new_first_sheet.cell(row = writing, column = 7).value = address2

#? What is carrier
def carrier(reading, writing):
    carrier = first_sheet.cell(row = reading, column = 12).value
    carrier_code = carrier_dict[carrier]
    new_first_sheet.cell(row = writing, column = 8).value = carrier_code

#Uses UPC code to get item num 
#Then uses item num to get item desc
def item_num_and_desc(reading, writing):
    upc_code = first_sheet.cell(row = reading, column = 4).value
    upc_code = str(upc_code)
    item_num = item_num_dict.get(upc_code, "!!! Update Item Sheet")
    item_num = str(item_num)
    new_first_sheet.cell(row = writing, column = 9).value = item_num
    #Use item num to get desc
    item_desc = description_dict.get(item_num, "!!! Update Item Sheet")
    new_first_sheet.cell(row = writing, column = 10).value = item_desc

def unit_price(reading, writing):
    unit_price = first_sheet.cell(row = reading, column = 16).value
    #Display 2 0's after decimal
    unit_price = float(unit_price)
    unit_price = "{:.2f}".format(unit_price)
    new_first_sheet.cell(row = writing, column = 11).value = unit_price

def quantity(reading, writing):
    quantity = first_sheet.cell(row = reading, column = 3).value
    new_first_sheet.cell(row = writing, column = 12).value = quantity

def terms(reading, writing):
    new_first_sheet.cell(row = writing, column = 14).value = 'NET 60'

#Use quantity and Unit price to calculate line total
def line_total():
    for row_num in range(2, new_first_sheet.max_row + 1):
        qty = new_first_sheet.cell(row = row_num, column = 11).value 
        price = new_first_sheet.cell(row = row_num, column = 12).value
        price = float(price)
        qty = float(qty)
        line_total = qty * price
        #Display 2 0's after decimal
        line_total = float(line_total)
        line_total = "{:.2f}".format(line_total)
        new_first_sheet.cell(row = row_num, column = 13).value = line_total

#Use po and line total to calculate total for each po num
def po_total():
    po_total = 0
    #first_new_po is the first po number of its kind, no duplicates
    first_new_po = new_first_sheet.cell(row = 2, column = 2).value
    #Initialize writing row to row 2
    po_total_writing_row = 2

    for row_num in range(2, new_first_sheet.max_row + 1):
        current_po = new_first_sheet.cell(row = row_num, column = 2).value
        line_total = new_first_sheet.cell(row = row_num, column = 13).value
        line_total = float(line_total)
        #print("first po is", first_new_po, "current po is", current_po, first_new_po == current_po)

        if first_new_po == current_po:
            po_total += line_total
       
        else: 
            #When po changes, write the po total to writing row
            new_first_sheet.cell(row = po_total_writing_row, column = 15).value = po_total
            #Reset po Total
            po_total = 0
            #Add current line total to new po total 
            po_total += line_total
            #Keep track of first new po to compare against current po
            first_new_po = new_first_sheet.cell(row = row_num, column = 2).value
            #Update writing row
            po_total_writing_row = row_num

    #At the end, write final po total once
    new_first_sheet.cell(row = po_total_writing_row, column = 15).value = po_total

def body():
    writing = 2
    for reading in range(8, first_sheet.max_row + 1):
        account(reading, writing)
        po_num(reading, writing)
        po_line(reading, writing)
        customer_name(reading, writing)
        address_1(reading, writing)
        phone_num(reading, writing)
        address_2(reading, writing)
        carrier(reading, writing)
        item_num_and_desc(reading, writing)
        unit_price(reading, writing)
        quantity(reading, writing)
        terms(reading, writing)
        writing += 1

def main():
    create_description_dict()
    create_carrier_dict()
    create_item_num_dict()
    titles()
    body()
    line_total()
    po_total()

main()

#Pass file name to save
new_wb.save("SBT_DICKS.xlsx")