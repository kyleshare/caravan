import openpyxl
import os

#MAC filepath
filepath = os.path.join('/Users', 'KyleShare', 'Programming', 'caravan', 'HomeDepot 01-07-19.xlsx' )

#WINDOWS filepath
#ilepath = os.path.join('C:\\', 'Users', 'CaravanArms', 'Desktop', 'HOMEDEPOT.XLSX' )


#Get workbook from filepath
wb = openpyxl.load_workbook(filepath)

#Get all sheetnames
sheet_names = wb.sheetnames

#Get first worksheet, since wb.active is set to 0 by default
first_sheet = wb.active

#Create new workbook
new_wb = openpyxl.Workbook()
new_first_sheet = new_wb.active


def titles():
    titles = ["ACCOUNT(SBT CODE)", "PO#", "PO LINE", "CUSTOMER NAME", "ADDRESS 1(2ND LINE)", \
    "PHONE# (3RD LINE)", "ADDRESS 2", "CARRIER", "ITEM#", "ITEM DESCRIPTION", \
    "UNIT PRICE", "QTY", "LINE TOTAL", "TERMS"]
    title_index = 0
    for column_num in range(1, 15):
        new_first_sheet.cell(row = 1, column = column_num).value = titles[title_index]
        title_index += 1

def account():
    new_row = 2
    for row_num in range(8, first_sheet.max_row + 1):
        new_first_sheet.cell(row = new_row, column = 1).value = 'HOMEDP'
        new_row += 1

def po_num():
    new_row = 2
    for row_num in range(8, first_sheet.max_row + 1):
        po = first_sheet.cell(row = row_num, column = 1).value
        new_first_sheet.cell(row = new_row, column = 2).value = po
        new_row += 1

def po_line():
    new_row = 2
    for row_num in range(8, first_sheet.max_row + 1):
        line = first_sheet.cell(row = row_num, column = 2).value
        new_first_sheet.cell(row = new_row, column = 3).value = line
        new_row += 1

def customer_name():
    new_row = 2
    for row_num in range(8, first_sheet.max_row + 1):
        name = first_sheet.cell(row = row_num, column = 6).value
        new_first_sheet.cell(row = new_row, column = 4).value = name
        new_row += 1

#<Address, N/A> or <Store #, Store Address>
def address_1():
    new_row = 2
    for row_num in range(8, first_sheet.max_row + 1):
        shipping1 = first_sheet.cell(row = row_num, column = 7).value
        shipping2 = first_sheet.cell(row = row_num, column = 8).value

        address1 = "{}, {}".format(shipping1, shipping2)
        new_first_sheet.cell(row = new_row, column = 5).value = address1
        new_row += 1

def phone_num():
    new_row = 2
    for row_num in range(8, first_sheet.max_row + 1):
        phone = first_sheet.cell(row = row_num, column = 15).value
        new_first_sheet.cell(row = new_row, column = 6).value = phone
        new_row += 1

#<City, State Zip>
#Initial 0's omitted on downloads, make Zip Code 5 digits
def address_2():
    new_row = 2
    for row_num in range(8, first_sheet.max_row + 1):
        city = first_sheet.cell(row = row_num, column = 9).value
        state = first_sheet.cell(row = row_num, column = 10).value
        #Zip comes in as ="#####", remove = and ""
        zip_code = first_sheet.cell(row = row_num, column = 11).value
        #zip_code = zip_code[2:7]
        #zip_code = str(zip_code)
        #zip_code = zip_code.zfill(5)

        address2 = "{}, {} {}".format(city, state, zip_code)
        new_first_sheet.cell(row = new_row, column = 7).value = address2
        new_row += 1

#? What is carrier
def carrier():
    new_row = 2
    for row_num in range(8, first_sheet.max_row + 1):
        new_first_sheet.cell(row = new_row, column = 8).value = 'Unknown'
        new_row += 1

def item_num():
    new_row = 2
    for row_num in range(8, first_sheet.max_row + 1):
        item_num = first_sheet.cell(row = row_num, column = 4).value
        new_first_sheet.cell(row = new_row, column = 9).value = item_num
        new_row += 1

#May have 2 item desc?
def item_desc():
    new_row = 2
    for row_num in range(8, first_sheet.max_row + 1):
        item_desc = first_sheet.cell(row = row_num, column = 5).value
        new_first_sheet.cell(row = new_row, column = 10).value = item_desc
        new_row += 1

def unit_price():
    new_row = 2
    for row_num in range(8, first_sheet.max_row + 1):
        price = first_sheet.cell(row = row_num, column = 16).value
        new_first_sheet.cell(row = new_row, column = 11).value = price
        new_row += 1

def quantity():
    new_row = 2
    for row_num in range(8, first_sheet.max_row + 1):
        quantity = first_sheet.cell(row = row_num, column = 3).value
        new_first_sheet.cell(row = new_row, column = 12).value = quantity
        new_row += 1

def terms():
    new_row = 2
    for row_num in range(8, first_sheet.max_row + 1):
        new_first_sheet.cell(row = new_row, column = 14).value = 'NET 60'
        new_row += 1

#Use quantity and Unit price to calculate line total
def line_total():
    for row_num in range(2, new_first_sheet.max_row + 1):
      line_total = new_first_sheet.cell(row = row_num, column = 11).value * \
      new_first_sheet.cell(row = row_num, column = 12).value
      new_first_sheet.cell(row = row_num, column = 13).value = line_total

def main():
    titles()
    account()
    po_num()
    po_line()
    customer_name()
    address_1()
    phone_num()
    address_2()
    carrier()
    item_num()
    item_desc()
    unit_price()
    quantity()
    terms()
    line_total()

main()

#Pass file name to save
new_wb.save("SBT_HOMEDEPOT.xlsx")