#import json
import openpyxl
import os

#If you have a Python object, you can convert it into a JSON string by using the json.dumps() method.
#obj = json.dumps()

#directory = os.getcwd()

#function takes file/folder names in path, returns correct file path based on os
filepath = os.path.join('/Users', 'KyleShare', 'Programming', 'caravan', 'Copy of PRODUCTION LOG _ 2019 _MASTER.xlsx' )

wb = openpyxl.load_workbook(filepath)
print(type(wb))


#Gets all sheetnames
sheet_names = wb.sheetnames
print("All sheet_names are", sheet_names)
#Gets first worksheet, since wb.active is set to 0 by default
first_sheet = wb.active
print("The first sheet is", first_sheet)

print(first_sheet.cell(row=1921, column=1).value)

#print(first_sheet.min_row)
#print(first_sheet.max_row)
#print(first_sheet.max_column)

#Tuple (Row number, Row dimension)
for row_num, j in first_sheet.row_dimensions.items():
    #Check if row is hidden
    if j.hidden == False:
        print("Row #", row_num)
        #ID is column 1 for every visible cell
        ID = (first_sheet.cell(row=row_num, column=1).value)
        #Status is column 1 for every visible cell
        status = (first_sheet.cell(row=row_num, column=3).value)
        #Filter out visible cells that contain no data
        if ID is not None and status is not None:
            print([ID, status])



'''
#Create workbook object, opens existing excel file to read
wb = openpyxl.Workbook()
print(wb)

print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name("Sheet")
print(sheet)

sheet["A1"].value == None
'''

'''
REQUIREMENTS FOR PROGRAM TO WORK
1) File path location must stay the same
2) Data must be on first sheet of excel workbook, currently called "2018"
    (It is okay to rename this sheet though)
3) Satus of item must be spelled correctly, 
    (But case does not matter)

Talking points for resume
Had to deal with deprecated warnings, using a feature that
will be removed in future versions of python.
This was an error siwht my get_sheet_names function
'''

