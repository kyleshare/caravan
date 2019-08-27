#Kyle Share 
import openpyxl
import os

#MAC filepath
#filepath = os.path.join('/Users', 'KyleShare', 'Programming', 'caravan', 'TRAGET.XLSX' )
#description_path = os.path.join('/Users', 'KyleShare', 'Programming', 'caravan', 'Item Description.xlsx')

#WINDOWS filepath
filepath = os.path.join('C:\\', 'Users', 'CaravanArms', 'Desktop', 'TARGET.XLSX' )
description_path = os.path.join('C:\\', 'Users', 'AKim', 'Desktop', 'Item Description.xlsx')

#Get workbook from filepath
wb = openpyxl.load_workbook(filepath)
description_wb = openpyxl.load_workbook(description_path)

#Get all sheetnames
sheet_names = wb.sheetnames
description_sheet_names = description_wb.sheetnames

#Get first worksheet, since wb.active is set to 0 by default
first_sheet = wb.active
description_sheet = description_wb.active

#Create new workbook
new_wb = openpyxl.Workbook()
new_first_sheet = new_wb.active

#Create dictionary of (ID: item description) pairs
#Make keys strings since key (item num) will also be string
#Can't be integers because some item numbers start with letters
def create_dict():
  description_list = []
  #Iterate through description excel sheet
  for row_num in range(2, description_sheet.max_row):
    key = str((description_sheet.cell(row=row_num, column=1).value))
    description = (description_sheet.cell(row=row_num, column=2).value)
    #Create tuples, append them to description list
    tupl = (key, description)
    description_list.append(tupl)
    #Make description dictionary global so item desctiption function can access it
    global description_dict
    description_dict = dict(description_list)
    
def titles():
    titles = ["ACCOUNT(SBT CODE)", "PO#", "PO LINE", "CUSTOMER NAME", "ADDRESS 1(2ND LINE)", \
    "PHONE# (3RD LINE)", "ADDRESS 2", "CARRIER", "ITEM#", "ITEM DESCRIPTION", \
    "UNIT PRICE", "QTY", "LINE TOTAL", "TERMS"]
    title_index = 0
    for column_num in range(1, 15):
        new_first_sheet.cell(row = 1, column = column_num).value = titles[title_index]
        title_index += 1

#Functions copy data from reading, write to writing
def account(writing, reading):
    new_first_sheet.cell(row = writing, column = 1).value = 'TARGET'

#Initial 0's omitted on downloads, make PO 10 digits
def po_num(writing, reading):
    po = first_sheet.cell(row = reading, column = 1).value
    po = str(po)
    po = po.zfill(10)
    new_first_sheet.cell(row = writing, column = 2).value = po

def po_line(writing, reading):
    line = first_sheet.cell(row = reading, column = 12).value
    new_first_sheet.cell(row = writing, column = 3).value = line

def customer_name(writing, reading):
    name = first_sheet.cell(row = reading, column = 62).value
    name = name.upper()
    new_first_sheet.cell(row = writing, column = 4).value = name


#<Street address, Appt/Suite>
#Appt/Suite may be in same cell as street address or 1 column right
def address_1(writing, reading):
    street_address = first_sheet.cell(row = reading, column = 63).value
    apartment = first_sheet.cell(row = reading, column = 64).value

    #if apartment exists on next column, add it to street address
    if apartment != None:
      street_address = "{} {}".format(street_address, apartment)

    street_address = street_address.upper()
    new_first_sheet.cell(row = writing, column = 5).value = street_address

def phone_num(writing, reading):
    phone = first_sheet.cell(row = reading, column = 52).value
    new_first_sheet.cell(row = writing, column = 6).value = phone

#<City, State Zip>
#Initial 0's omitted on downloads, make Zip Code 5 digits
def address_2(writing, reading):
    city = first_sheet.cell(row = reading, column = 65).value
    state = first_sheet.cell(row = reading, column = 66).value
    zip_code = first_sheet.cell(row = reading, column = 67).value
    zip_code = str(zip_code)
    zip_code = zip_code.zfill(5)

    address2 = "{}, {} {}".format(city, state, zip_code)
    address2 = address2.upper()
    new_first_sheet.cell(row = writing, column = 7).value = address2

def carrier(writing, reading):
    new_first_sheet.cell(row = writing, column = 8).value = '3PT FDXG'

def item_num(writing, reading):
    item_num = first_sheet.cell(row = reading, column = 97).value
    new_first_sheet.cell(row = writing, column = 9).value = item_num

#Use item_num as key in dictionary to get item desc
def item_desc(writing, reading):
    item_num = first_sheet.cell(row = reading, column = 97).value
    item_num = str(item_num)
    item_desc = description_dict[item_num]
    new_first_sheet.cell(row = writing, column = 10).value = item_desc

def unit_price(writing, reading):
    unit_price = first_sheet.cell(row = reading, column = 15).value
    #Display 2 0's after decimal
    unit_price = float(unit_price)
    unit_price = "{:.2f}".format(unit_price)
    new_first_sheet.cell(row = writing, column = 11).value = unit_price

def quantity(writing, reading):
    quantity = first_sheet.cell(row = reading, column = 13).value
    new_first_sheet.cell(row = writing, column = 12).value = quantity

def terms(writing, reading):
    new_first_sheet.cell(row = writing, column = 14).value = 'NET 30'

def main_info(writing, reading):
  account(writing, reading)
  po_num(writing, reading)
  customer_name(writing, reading)
  address_1(writing, reading)
  phone_num(writing, reading)
  address_2(writing, reading)
  carrier(writing, reading)
  terms(writing, reading)

def details(writing, reading):
  po_line(writing, reading)
  item_num(writing, reading)
  item_desc(writing, reading)
  unit_price(writing, reading)
  quantity(writing, reading)

def body():
    reading = 2
    writing = 2
    first_po =  first_sheet.cell(row = reading, column = 1).value
    first_po_row = 2

    for row in range(first_sheet.max_row):
        #Iterate until run out of PO's. (For loop range is too large )
        if first_po != None:

          #Save main info since it is reused for multiple orders with same PO
          main_info(writing, reading)
          main_info_reading = reading

          #Get details by reading from next row 
          reading += 1
          details(writing, reading)

          #After writing main info and details, write to and read from new row
          #Skip extra reading line, Target downloads has an extra line w/ no info
          reading += 2
          writing += 1

          #Save current PO to check against previous
          current_po = first_sheet.cell(row = reading, column = 1).value

          #While PO stays same, create new lines to write to but pull main info from main line
          #***This doesn't happen for target, but it does for other companies***
          while current_po == first_po:
            #Copy new details for each line, but same main info
            main_info(writing, first_po_row)
            details(writing, reading)

            reading += 1
            writing += 1
            current_po = first_sheet.cell(row = reading, column = 1).value

          #When PO changes, save new PO and revert to normal pattern
          else:
            first_po = first_sheet.cell(row = reading, column = 1).value
            first_po_row = first_sheet.cell(row = reading, column = 1).row

#Use quantity and Unit price to calculate line total
def line_total():
    for row_num in range(2, new_first_sheet.max_row + 1):
      qty =  new_first_sheet.cell(row = row_num, column = 11).value
      price = new_first_sheet.cell(row = row_num, column = 12).value
      price = float(price)
      qty = float(qty)
      line_total = qty * price
      #Display 2 0's after decimal
      line_total = "{:.2f}".format(line_total)
      new_first_sheet.cell(row = row_num, column = 13).value = line_total

def main():
    create_dict()
    titles()
    body()
    line_total()

main()

new_wb.save("SBT_TARGET.xlsx")