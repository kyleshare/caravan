#Kyle Share
import openpyxl
import os

#MAC filepath
#filepath = os.path.join('/Users', 'KyleShare', 'Programming', 'caravan', '2WALMART W.H.XLSX' )
#description_path = os.path.join('/Users', 'KyleShare', 'Programming', 'caravan', 'Item Description.xlsx')

#WINDOWS filepath
filepath = os.path.join('C:\\', 'Users', 'Akim', 'Desktop', 'WALMART.xlsx' )
description_path = os.path.join('C:\\', 'Users', 'AKim', 'Desktop', 'Item Description.xlsx')

#Get workbook from filepath
wb = openpyxl.load_workbook(filepath)
description_wb = openpyxl.load_workbook(description_path)

#Get all sheetnames
sheet_names = wb.sheetnames
description_sheet_names = description_wb.sheetnames

#Get first worksheet, since wb.active is set to 0 by default
first_sheet = wb.active
description_sheet = description_wb.active

#Create new workbook
new_wb = openpyxl.Workbook()
new_first_sheet = new_wb.active

#Create dictionary of (ID: item description) pairs
def create_description_dict():
    description_list = []
    #Iterate through description excel sheet
    for row_num in range(2, description_sheet.max_row):
        key = str((description_sheet.cell(row=row_num, column=1).value))
        description = (description_sheet.cell(row=row_num, column=2).value)
        #Create tuples, append them to description list
        tupl = (key, description)
        description_list.append(tupl)
        description_dict = dict(description_list)
    return description_dict

def create_price_dict():
    price_list = []
    #Iterate through prices excel sheet
    for row_num in range(2, description_sheet.max_row):
        key = str((description_sheet.cell(row=row_num, column=1).value))
        #Walmart price is column 3
        price = (description_sheet.cell(row=row_num, column=5).value)
        #Create tuples, append them to description list
        tupl = (key, price)
        price_list.append(tupl)
        price_dict = dict(price_list)
    return price_dict

def create_carrier_dict():
    carrier_list = []
    #Iterate through prices excel sheet
    for row_num in range(2, description_sheet.max_row):
        key = str((description_sheet.cell(row=row_num, column=3).value))
        #Walmart price is column 3
        carrier = (description_sheet.cell(row=row_num, column=4).value)
        #Create tuples, append them to description list
        tupl = (key, carrier)
        carrier_list.append(tupl)
        carrier_dict = dict(carrier_list)
    return carrier_dict


def titles():
    titles = ["ACCOUNT(SBT CODE)", "PO#", "PO LINE", "CUSTOMER NAME", "ADDRESS 1(2ND LINE)", \
    "PHONE# (3RD LINE)", "ADDRESS 2", "CARRIER", "ITEM#", "ITEM DESCRIPTION", \
    "UNIT PRICE", "QTY", "LINE TOTAL", "TERMS"]
    title_index = 0
    for column_num in range(1, 15):
        new_first_sheet.cell(row = 1, column = column_num).value = titles[title_index]
        title_index += 1

def account():
    for row_num in range(2, first_sheet.max_row + 1):
        new_first_sheet.cell(row = row_num, column = 1).value = 'WMECOM'

#Initial 0's omitted on downloads, make PO 10 digits
def po_num():
    for row_num in range(2, first_sheet.max_row + 1):
        po = first_sheet.cell(row = row_num, column = 1).value
        po = str(po)
        po = po.zfill(10)
        new_first_sheet.cell(row = row_num, column = 2).value = po

def po_line():
    for row_num in range(2, first_sheet.max_row + 1):
        line = first_sheet.cell(row = row_num, column = 16).value
        new_first_sheet.cell(row = row_num, column = 3).value = line

#Maybe I should hide this for legal reasons?
def customer_name():
    for row_num in range(2, first_sheet.max_row + 1):
        name = first_sheet.cell(row = row_num, column = 5).value
        name = name.upper()
        new_first_sheet.cell(row = row_num, column = 4).value = name

def address_1():
    for row_num in range(2, first_sheet.max_row + 1):
        address = first_sheet.cell(row = row_num, column = 9).value
        address = address.upper()
        new_first_sheet.cell(row = row_num, column = 5).value = address

def phone_num():
    for row_num in range(2, first_sheet.max_row + 1):
        phone = first_sheet.cell(row = row_num, column = 7).value
        new_first_sheet.cell(row = row_num, column = 6).value = phone

#<City, State Zip>
#Initial 0's omitted on downloads, make Zip Code 5 digits
def address_2():
    for row_num in range(2, first_sheet.max_row + 1):
        city = first_sheet.cell(row = row_num, column = 11).value
        state = first_sheet.cell(row = row_num, column = 12).value
        zip_code = first_sheet.cell(row = row_num, column = 13).value
        zip_code = str(zip_code)
        zip_code = zip_code.zfill(5)

        address2 = "{}, {} {}".format(city, state, zip_code)
        address2 = address2.upper()
        new_first_sheet.cell(row = row_num, column = 7).value = address2

def carrier(carrier_dict):
    for row_num in range(2, first_sheet.max_row + 1):
        carrier = first_sheet.cell(row = row_num, column = 24).value
        carrier_code = carrier_dict[carrier]
        new_first_sheet.cell(row = row_num, column = 8).value = carrier_code

def item_num():
    for row_num in range(2, first_sheet.max_row + 1):
        item_num = first_sheet.cell(row = row_num, column = 18).value
        new_first_sheet.cell(row = row_num, column = 9).value = item_num

#May have 2 item desc
def item_desc(description_dict):
    for row_num in range(2, first_sheet.max_row + 1):
        item_num = first_sheet.cell(row = row_num, column = 18).value
        item_desc = description_dict[item_num]
        new_first_sheet.cell(row = row_num, column = 10).value = item_desc

def unit_price(price_dict):
    for row_num in range(2, first_sheet.max_row + 1):
        item_num = first_sheet.cell(row = row_num, column = 18).value
        unit_price = price_dict[item_num]

        #Display 2 0's after decimal
        unit_price = float(unit_price)
        unit_price = "{:.2f}".format(unit_price)

        new_first_sheet.cell(row = row_num, column = 11).value = unit_price

#Walmart stores quantity as text, convert to int
def quantity():
    for row_num in range(2, first_sheet.max_row + 1):
        quantity = first_sheet.cell(row = row_num, column = 21).value
        quantity = int(quantity)
        new_first_sheet.cell(row = row_num, column = 12).value = quantity


def terms():
    for row_num in range(2, first_sheet.max_row + 1):
        new_first_sheet.cell(row = row_num, column = 14).value = 'NET 30'

#Use quantity and Unit price to calculate line total
def line_total():
    for row_num in range(2, new_first_sheet.max_row + 1):
            qty = new_first_sheet.cell(row = row_num, column = 11).value 
            price = new_first_sheet.cell(row = row_num, column = 12).value
            price = float(price)
            qty = float(qty)
            line_total = qty * price
            #Display 2 0's after decimal
            line_total = float(line_total)
            line_total = "{:.2f}".format(line_total)
            new_first_sheet.cell(row = row_num, column = 13).value = line_total

#Fixes name for orders that ship to store, 
def fix_name():
    for row_num in range(2, first_sheet.max_row + 1):
        #Check if order has store id
        if first_sheet.cell(row = row_num, column = 8).value:
            name = first_sheet.cell(row = row_num, column = 10).value
            name = name.upper()
            store_id = first_sheet.cell(row = row_num, column = 8).value

            fixed_name = "{} STORE ID: {}".format(name, store_id)
            new_first_sheet.cell(row = row_num, column = 4).value = fixed_name


def main():
    description_dict = create_description_dict()
    price_dict = create_price_dict()
    carrier_dict = create_carrier_dict()
    titles()
    account()
    po_num()
    po_line()
    customer_name()
    address_1()
    phone_num()
    address_2()
    carrier(carrier_dict)
    item_num()
    item_desc(description_dict)
    unit_price(price_dict)
    quantity()
    terms()
    #Implement after unit price is fixed
    line_total()
    fix_name()

main()

#Pass file name to save
new_wb.save("SBT_WALMART.xlsx")
